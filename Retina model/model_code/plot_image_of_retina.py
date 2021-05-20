import sys
import numpy as np
import os
import pylab as pl
import seaborn as sns
sns.set()
#    import matplotlib.pyplot as mt
# sys.path.append(r"C:\Users\Bob\engert_lab\free_swimming_4fish_setup\modules")
#
# from shared import Shared
from Eye_model_function_curve import *
import pandas as pd
from openpyxl import load_workbook
from fitEllipse import fit_ellipse

# dddddd dddd
# we need to define the speed profile of the moving dot, and sample speeds
# from it to use for stimulus presentations.

# lets use sigmoid up and sigmid down - i.e. symmetric distribution
# Fs = 50
# xsig = np.arange(0.7)/Fs
# x0 = 3.5/Fs
# maxV = 4.25  # maximum speed (a bit higher since the sigmid will only approach maxV
# slope = 41  # set the slope of response
# Sup = maxV/(1+np.exp(-slope*(xsig-x0)))
# Sdown = maxV/(1+np.exp(slope*(xsig-x0)))

####VARIABLE ON TIME OF STIMULUS:
dt = 1 / 90
ISI = 2.8  # sec before and after the stim
loom_time = 5  # sec

###VARIABLE ON SIZE OF ARENA AND STIMULUS:
arena_radius = 12.6 / 2  # cm - full size
radius = 1  # radius of the dot
edges = 30  # number of edges of the polygon (dot)
angle_to_fish = 70  # angle to Oy
angle_ver = 70 # angle to Oy 
sizes = 0.075 / arena_radius  # REAL SIZE OF DOT IN RESPECT TO ARENA SIZE
angs = np.linspace(0, 360, edges)
list_vertex = []
list_vertex0 = []
minor_p = major_p = sizes
ratio = 1

###VARIABLE ON POSITION OF STIMULUS:
distance_xy_right = 0.9  # in cm
distance_xy_left = 0.9  # in cm
dist_to_border = distance_xy_right / arena_radius - sizes
dist_to_center = distance_xy_right/arena_radius

###VARIABLE ON PROPERTIES OF EYE:
dist_btw_eye = 0.12 / arena_radius  # distance between the two eyes
x_leye = -0.06 / arena_radius  # x position of the left eye
y_leye = 0  # y position of the left eye
h_eye = 0.5 / arena_radius  # the z position of the fish from the ground
eye_rad = 0.045 / arena_radius  # fish eye's radius
angle_retina = 26.5 #degree
x_reye = x_leye + dist_btw_eye  # x position of  the right eye
y_reye = y_leye  # y position of the right eye
z_reye = z_leye = h_eye  # the z position of the fish from the ground

vector_retina_right = np.array(
    [np.cos(np.radians(26.5)), -np.sin(np.radians(26.5)), 0])  # normal vector of the right retina plane
vector_retina_left = np.array(
    [-np.cos(np.radians(26.5)), -np.sin(np.radians(26.5)), 0])  # normal vector of the left retina plane
list_of_input = [x_reye, y_reye, z_reye, x_leye, y_leye, z_leye, vector_retina_right, vector_retina_left]
properties = [eye_rad, h_eye]


x_center_retina_right = x_reye - eye_rad*np.cos(np.radians(angle_retina))
y_center_retina_right = y_reye - eye_rad*np.sin(np.radians(angle_retina))
print("y center retina:", y_center_retina_right * arena_radius)
x_center_retina_left = x_leye + eye_rad*np.cos(np.radians(angle_retina))
y_center_retina_left = y_center_retina_right
print("x ret:",x_center_retina_right)
####################


def squeeze_point_along_axis(point, scale_factor):  # squeeze a circle to an ellipse along an axis
    return (point[0], point[1] * scale_factor)


def rotate_point_around_origin(coor, angle):  # transform between two coordinate systems rotated by an angle
    angle = np.radians(angle)
    x, y = coor
    xx = x * np.cos(angle) + y * np.sin(angle)
    yy = -x * np.sin(angle) + y * np.cos(angle)

    return xx, yy


class imageToDot(image_computation):
    ## compute vertices position for a normal circle dot
    for j in range(len(angs)):
        ang0 = angs[j]
        list_vertex.append([np.cos(ang0 * np.pi / 180) * radius, np.sin(ang0 * np.pi / 180) * radius])
    for j in range(len(angs)):
        ang0 = angs[j]
        list_vertex0.append([np.cos(ang0 * np.pi / 180) * 0, np.sin(ang0 * np.pi / 180) * 0])

    def compute_image(self, list_of_input, eye_properties):
        eye_rad, h_eye = eye_properties

        x_right = (sizes + dist_to_border) * np.sin(angle_to_fish * np.pi / 180.)
        y_right = (sizes + dist_to_border) * np.cos(angle_to_fish * np.pi / 180.)

        x_left = (sizes + dist_to_border) * np.sin(-angle_to_fish * np.pi / 180.)
        y_left = (sizes + dist_to_border) * np.cos(-angle_to_fish * np.pi / 180.)

        x_pos = [x_right, x_left]
        y_pos = [y_right, y_left]

        points = {}

        for i in range(2):
            points[i] = [np.array([(vertex[1] * sizes + y_pos[i]) for vertex in list_vertex]),  # y
                         np.array([(vertex[0] * sizes + x_pos[i]) for vertex in list_vertex]),  # x
                         np.zeros((len(list_vertex)))]

        x_new_right, y_new_right, z_new_right, x_new_left, y_new_left, z_new_left = self.find_new_coordinate_list(
            points, 2, list_of_input, eye_rad)

        x_new = [x_new_right[0], x_new_left[1]]
        y_new = [y_new_right[0], y_new_left[1]]
        z_new = [z_new_right[0], z_new_left[1]]
        print("z_new:",z_new)
        return x_new, y_new, z_new, x_pos, y_pos, points

    def traceToDot(self, list_of_input, eye_properties, list_vertex, sizes):
        x_new, y_new, z_new, x_pos, y_pos, points = self.compute_image(list_of_input, eye_properties)

        x_reye, y_reye, z_reye, x_leye, y_leye, z_leye, _, _ = list_of_input
        x_ = [x_reye, x_leye]
        y_ = [y_reye, y_leye]
        z_ = [z_reye, z_leye]

        eye_rad, h_eye = eye_properties
        stimuli = [[list_vertex0], [list_vertex0]]
        raw_stimuli =  [[], []]
        phi_delta = 3
        eye_rad_zoom = 10
        dx_dy_twoSides = [[], []]
        stimuli_amount = 4
        normal_fact = 0
        image_position = [[],[]]
        for i in range(2):  # compute next images for both eyes and trace back to the dot on both eyes

            x_new_ = np.array(x_new[i])
            y_new_ = np.array(y_new[i])
            z_new_ = np.array(z_new[i])
            #image_position[i].append([x_new_, y_new_, z_new_])
            
            phi_origin = np.arcsin((z_new_ - z_[i]) / eye_rad)

            theta_origin = np.arccos((x_new_ - x_[i]) / (eye_rad * np.cos(phi_origin)))

            theta_mean = np.median(theta_origin)

            y_max = 0
            x_max = 0
            rr_list = []
            cc_list = []
            
            
            for j in range(stimuli_amount):
                # try find the polar (phi) and the azimuth angle (theta) of the image:

                ##for minh: if theta is updated and phi is not --> image elevate on the eye. if phi is updated while theta_mean is used --> rotate around the eye.
                phi_new = phi_origin - np.radians(phi_delta) #phi is not updated
                theta_new = (theta_origin - theta_mean) * np.cos(phi_origin) / np.cos(phi_new) + theta_mean

                # find coordinate of the next image on retina
                z_new2 = z_[i] + eye_rad * np.sin(phi_new)
                x_new2 = x_[i] + eye_rad * np.cos(phi_new) * np.cos(theta_new)
                y_new2 = y_[i] - eye_rad * np.cos(phi_new) * np.sin(theta_new)
                image_position[i].append([x_new_, y_new_, z_new_])  # image position of previous image
                # Trace back to find the real dots and plot them

                dots = self.dot_find(x_new_, y_new_, z_new_, x_[i], y_[i], z_[i], eye_rad)



                x_mean = np.mean(dots[1])
                y_mean = np.mean(dots[0])
                dx_dy_twoSides[i].append([np.sqrt(x_mean ** 2 + y_mean ** 2)]) ### this is the next position of the stimulus

                rr = [(j - y_mean) for j in dots[0]]  # y
                cc = [(k - x_mean) for k in dots[1]] # x

                if i == 0 and j == 0:
                    normal_fact = max(max(rr), max(cc))


                if len(rr) > 0:
                    if max(rr) > y_max:
                        y_max = max(abs(np.array(rr)))
                    if max(cc) > x_max:
                        x_max = max(abs(np.array(cc)))
                rr_list.append(rr)
                cc_list.append(cc)

                z_new_ = z_new2
                x_new_ = x_new2
                y_new_ = y_new2
                phi_origin = phi_new
                theta_origin = theta_new

                vertices = [[cc_list[j][i] / normal_fact, rr_list[j][i] / normal_fact] for i in range(len(rr_list[j]))]
                stimuli[i].append(vertices)
                raw_vertices = [[dots[1][i] , dots[0][i] ] for i in range(len(dots[0]))] ### raw vertices of the last image's stimulus
                raw_stimuli[i].append(raw_vertices)

        print("len stimuli:", len(stimuli))
        return stimuli, dx_dy_twoSides, normal_fact, points, raw_stimuli, image_position

    def mix_two_list(self, list1, list2, ISI, loom_time):
        list_mix = []
        t_stimuli = []
        for i in range(len(list1)):
            for j in range(len(list2)):
                if i != j:
                    list_mix.append([list1[i], list2[j]])
                    t_stimuli.append(2 * ISI + loom_time)
        return list_mix, t_stimuli


image_reverse = imageToDot()
stimuli, dx_dy_twoSides, scale, points, raw_stimuli, image_position = image_reverse.traceToDot(list_of_input, properties, list_vertex, sizes)
print("len stimuli one eye:", len(stimuli[0]))
print("len stimuli two eye combined:", len(stimuli))
#stimuli, t_stimuli = image_reverse.mix_two_list(stimuli[0], stimuli[1], ISI, loom_time)
#raw_stimuli, _ = image_reverse.mix_two_list(raw_stimuli[0], raw_stimuli[1], ISI, loom_time)
#dx_dy_twoSides, _ = image_reverse.mix_two_list(dx_dy_twoSides[0], dx_dy_twoSides[1], ISI, loom_time)
print("scale:",scale)
print("dot on right side:", points[0])
####Suppose the dot is at x,y from the fish

print(len(stimuli[0]), len(raw_stimuli[0]))
print("trial length", 2 * ISI + loom_time)
print("distance_to_fish:", dx_dy_twoSides)
# stimuli = [0]

#####This is to plot patch to see how the dots look like:



from matplotlib.collections import PatchCollection
import matplotlib.pyplot as plt
import matplotlib as mpl
import math

def sort_vertex(vertex_list):
    cent = (sum([p[0] for p in vertex_list]) / len(vertex_list),
            sum([p[1] for p in vertex_list]) / len(vertex_list))
    # sort by polar angle
    vertex_list.sort(key=lambda p: math.atan2(p[1] - cent[1], p[0] - cent[0]))
    return vertex_list


print(len(stimuli))
# for i in range(len(raw_stimuli[0])):
#     vertices_right = raw_stimuli[0][i]
#     print("pre sort:", vertices_right)
    
#     vertices_right = sort_vertex(vertices_right)
#     print("post sort:", vertices_right)
    
#     polygon = mpl.patches.Polygon(np.array(vertices_right))
#     patches.append(polygon)
# p = PatchCollection(patches, cmap=mpl.cm.jet, alpha=0.5)
# colors = 100 * np.random.rand(len(patches))
# p.set_array(np.array(colors))
# ax0.set_xlim(-1,1)
# ax0.set_ylim(-1,1)
# ax0.add_collection(p)
# plt.gca().set_aspect('equal', adjustable='box')
# plt.show()
# fig1, ax1 = plt.subplots(1,1)
# ax1.scatter(points[0][1], points[0][0])
# plt.show()

print("so hinh anh:", len(image_position[0]))


####PLOT EYE ======================================================================================================================

u, v = np.mgrid[np.radians(55):np.radians(197+55):1000j, 0:np.pi:1000j]
x = x_reye + eye_rad * np.cos(u)*np.sin(v)
y = y_reye + eye_rad * np.sin(u)*np.sin(v)
z = z_reye+ eye_rad * np.cos(v)
xl = x_leye + eye_rad * np.cos(u)*np.sin(v)
yl= y_leye + eye_rad * np.sin(u)*np.sin(v)
zl = z_leye+ eye_rad * np.cos(v)

fig = plt.figure()
ax2 = fig.add_subplot(122, projection='3d')
ax1 = fig.add_subplot(121)
#camera = Camera(fig)

colors = ['tab:red','tab:orange','gold','springgreen','lime','c', 'b','m','k']

for i in range(len(image_position[0])):
    patches = []
    vertices_right = raw_stimuli[0][i]
    polygon = mpl.patches.Polygon(np.array(vertices_right))
    patches.append(polygon)
    p = PatchCollection(patches, cmap=mpl.cm.jet, alpha=0.5, match_original=True, facecolor = colors[i], edgecolors = colors[i])
    #p.set_array(colors[i])
    ax1.set_xlim(-1,1)
    ax1.set_ylim(-1,1)
    ax1.add_collection(p)
    plt.gca().set_aspect('equal', adjustable='box')
    ax1.scatter(0,0, marker = 'o', s = 15, color = 'r')
    ax1.grid(False)
    ax2.plot_wireframe(x, y, z, color="lightblue")
    ax2.scatter(image_position[0][i][0], image_position[0][i][1], image_position[0][i][2], marker = 'o', facecolor = colors[i])
    ax2.scatter(x_center_retina_right, y_center_retina_right, z_reye, marker='o', color = 'r')
    # ax2.set_xlabel('Left ------> Right')
    # ax2.set_ylabel('Front ------> Back')
    # ax2.set_zlabel('Above ------> Below')
    ax2.grid(False)
    ax2.set_xticklabels([])
    ax2.set_yticklabels([])
    ax2.set_zticklabels([])
    ax2.set_xticks([])
    ax2.set_yticks([])
    ax2.set_zticks([])
    #camera.snap()

#animation = camera.animate(interval=30)
plt.show()


####================================================================================================================================
from openpyxl import Workbook
filename = r'/Users/nguyetnguyen/Downloads/Retina model/Retina model/model_code/similar_image_data.xlsx'

book = Workbook()
book.save(filename = filename)
book = load_workbook(filename)
writer = pd.ExcelWriter(filename, engine = 'openpyxl')
writer.book = book


filename2 = r'/Users/nguyetnguyen/Downloads/Retina model/Retina model/model_code/similar_image_properties.xlsx'


book2 = Workbook()
book2.save(filename = filename2)
book2 = load_workbook(filename2)
writer2 = pd.ExcelWriter(filename2, engine = 'openpyxl')
writer2.book = book2

for i in range(len(image_position[0])):
    tmp = []
    df = pd.DataFrame.empty
    df2 = pd.DataFrame.empty
    vertices_right = np.asarray(raw_stimuli[0][i]).T
    
    x_point = vertices_right[0,:]
    y_point = vertices_right[1,:]
    major_axis, minor_axis, _, _, _ = fit_ellipse(x_point, y_point)
    aspect_ratio = major_axis/minor_axis
    
    print("x_point:", x_point)
    z_point = np.zeros((len(x_point)))
    arena_radius_array = np.ones(len(x_point)) 
    
    distance_to_center =  dx_dy_twoSides[0][i][0] * arena_radius
    print("z_point:", z_point)
    tmp = list(zip(x_point, y_point, z_point, image_position[0][i][0] , image_position[0][i][1] , image_position[0][i][2] ))
    
    
    df = pd.DataFrame(tmp)
    df = df * arena_radius
    #print(df_right)
    
    if i == 0:
        std= book.get_sheet_by_name('Sheet')
        book.remove_sheet(std)
        std2= book2.get_sheet_by_name('Sheet')
        book2.remove_sheet(std2)
    df.to_excel(writer, index = False, header = ['x_point',
                                                'y_point',
                                                'z_point',
                                                'x_image_right',
                                                'y_image_right',
                                                'z_image_right',], sheet_name= 'image_' + str(i))
    df2 = pd.DataFrame(np.array([[major_axis * arena_radius, minor_axis * arena_radius, aspect_ratio, angle_to_fish, distance_to_center]]))
    df2.to_excel(writer2, index = False, header = ['major axis', 'minor axis', 'aspect ratio','angle to fish', 'distance to center'], sheet_name = 'image_' + str(i))
    filename3= r'/Users/nguyetnguyen/Downloads/Retina model/Retina model/model_code/similar_image_parameters.xlsx'
    df3 = pd.DataFrame(np.array([[ratio, minor_p* arena_radius, major_p * arena_radius, dist_to_center * arena_radius, angle_to_fish, angle_ver
                                  , x_reye * arena_radius, y_reye* arena_radius, z_reye* arena_radius, x_center_retina_right *arena_radius ,
                                  y_center_retina_right * arena_radius, eye_rad* arena_radius]]))
    
    df3.to_excel(filename3,  index = False,
                                         header = ['ratio', 'minor_p', 'major_p', 'dist_to_center', 'angle_to_fish', 'angle_ver',
                                                   'x_eye_ball_right', 'y_eye_ball_right', 'z_eye_ball_right', 'x_center_retina_right', 'y_center_retina_right', 'eye_rad'])
    
         
writer.save()
writer.close()
writer2.save()
writer2.close()